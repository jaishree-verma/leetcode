import openpyxl as xl
# to add chart -> package = openpyxl, module = chart,   = Barchart and Refernce
from openpyxl.chart import BarChart , Reference
wb = xl.load_workbook(r'C:\leetcode\Python\Project1\transaction.xlsx')          # file name 
sheet = wb['Sheet1']
cell = sheet['a1']
# print(cell.value)
# print("_________")
# output : Transaction
# print(sheet.max_row)      # to get max rows = 4
# print("_________")
for row in range(2, sheet.max_row + 1):   # iterate through each row to get column c values
    # print(row)
    cell = sheet.cell(row, 3)            # row,column = 3
    corrected_price = cell.value * 0.9       # multiply that column with 0.9
    corrected_price_cell = sheet.cell(row,4) # now we add 4th column and store there in col 4 so that col 4 has value with formula ans 
    corrected_price_cell.value = corrected_price
    # print(cell.value)                        # we got values from 3rd column

# we use refernce class to set range of values
# we select sheet and select cells in row 2 to max row (4) available in that sheet
# we want values in 4th column only
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

# create instance of Barchart Class
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('transaction2.xlsx')
# we create and save it in a new file
# this creates another column with those calculated values





# clean code 

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

    chart = BarChart() 
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')
    wb.save(filename)
